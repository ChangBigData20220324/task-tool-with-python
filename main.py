import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
import sqlite3
from datetime import datetime, timedelta
import pytz
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class TaskToolGUI:
    def __init__(self, master):
        self.master = master
        master.title("Daily Task Recorder")
        master.geometry("1000x600")

        # Create notebook for multiple pages
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Page 1: Input and Weekly View
        self.page1 = ttk.Frame(self.notebook)
        self.notebook.add(self.page1, text="Input & Weekly View")

        # Page 2: Excel Export
        self.page2 = ttk.Frame(self.notebook)
        self.notebook.add(self.page2, text="Excel Export")

        self.setup_page1()
        self.setup_page2()

        # Initialize database
        self.init_db()

    def setup_page1(self):
        # Left side: Input fields
        input_frame = ttk.Frame(self.page1)
        input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        today = datetime.now().date()
        days_until_thursday = (3 - today.weekday()) % 7
        thursday = today + timedelta(days=days_until_thursday)
        
        ttk.Label(input_frame, text="Settlement Date:").grid(row=0, column=0, sticky="w", pady=5)
        self.date_entry = DateEntry(input_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.date_entry.set_date(thursday)
        self.date_entry.grid(row=0, column=1, sticky="w", pady=5)

        ttk.Label(input_frame, text="Name:").grid(row=1, column=0, sticky="w", pady=5)
        self.name_entry = ttk.Entry(input_frame)
        self.name_entry.grid(row=1, column=1, sticky="we", pady=5)

        ttk.Label(input_frame, text="Task Content:").grid(row=2, column=0, sticky="w", pady=5)
        self.task_content = tk.Text(input_frame, height=5)
        self.task_content.grid(row=2, column=1, sticky="we", pady=5)

        ttk.Label(input_frame, text="Score:").grid(row=3, column=0, sticky="w", pady=5)
        self.score_entry = ttk.Entry(input_frame)
        self.score_entry.insert(0, "1")  # Set default value to "1"
        self.score_entry.grid(row=3, column=1, sticky="we", pady=5)

        ttk.Button(input_frame, text="Add Task", command=self.add_task).grid(row=4, column=0, columnspan=2, pady=10)

        # Right side: Weekly View
        weekly_frame = ttk.Frame(self.page1)
        weekly_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        ttk.Label(weekly_frame, text="Weekly View").pack()
        self.weekly_view = ttk.Treeview(weekly_frame, columns=("Date", "Name", "Task", "Score"), show="headings")
        self.weekly_view.heading("Date", text="Date")
        self.weekly_view.heading("Name", text="Name")
        self.weekly_view.heading("Task", text="Task")
        self.weekly_view.heading("Score", text="Score")
        self.weekly_view.pack(fill=tk.BOTH, expand=True)

        ttk.Button(weekly_frame, text="Refresh Weekly View", command=self.update_weekly_view).pack(pady=10)

    def setup_page2(self):
        ttk.Button(self.page2, text="Check Weekly Scores", command=self.check_weekly_scores).pack(pady=10)
        ttk.Button(self.page2, text="Export to Excel", command=self.export_to_excel).pack(pady=10)


    def init_db(self):
        conn = sqlite3.connect('tasks.db')
        c = conn.cursor()
        # ... existing code ...

        c.execute('''CREATE VIEW IF NOT EXISTS WEEKTASK_REPORT AS 
                    SELECT date, task, name,
                    CASE WHEN name = 'A' THEN score ELSE 0 END 'A分數',
                    CASE WHEN name = 'B' THEN score ELSE 0 END 'B分數',
                    CASE WHEN name = 'C' THEN score ELSE 0 END 'C分數',
                    strftime('%Y%m', date) YYYYMM,
                    strftime('%Y', date) YYYY
                    FROM tasks''')

        conn.commit()
        conn.close()


    def add_task(self):
        local_tz = pytz.timezone('Asia/Taipei')
        local_time = datetime.now(local_tz)
        
        conn = sqlite3.connect('tasks.db')
        c = conn.cursor()
        c.execute("""INSERT INTO tasks (date, name, task, score, created_at)
                    VALUES (?,?,?,?,?)""", (
            self.date_entry.get_date().strftime('%Y-%m-%d'),
            self.name_entry.get(),
            self.task_content.get("1.0", tk.END).strip(),
            self.score_entry.get(),
            local_time.strftime('%Y-%m-%d %H:%M:%S')
        ))
        conn.commit()
        conn.close()
        self.update_weekly_view()

    def update_weekly_view(self):
        for i in self.weekly_view.get_children():
            self.weekly_view.delete(i)

        conn = sqlite3.connect('tasks.db')
        c = conn.cursor()
        
        today = datetime.now().date()
        days_until_thursday = (3 - today.weekday()) % 7
        thursday = today + timedelta(days=days_until_thursday)
        
        c.execute("SELECT date, name, task, score FROM tasks WHERE date = ?", 
                (thursday.strftime('%Y-%m-%d'),))
        for row in c.fetchall():
            self.weekly_view.insert("", "end", values=row)
        conn.close()


    def check_weekly_scores(self):
        conn = sqlite3.connect('tasks.db')
        c = conn.cursor()
        
        today = datetime.now().date()
        days_until_thursday = (3 - today.weekday()) % 7
        thursday = today + timedelta(days=days_until_thursday)
        
        c.execute("SELECT name, SUM(score) FROM tasks WHERE date = ? GROUP BY name", 
                (thursday.strftime('%Y-%m-%d'),))
        results = c.fetchall()
        conn.close()

        # Display results in a new window
        result_window = tk.Toplevel(self.master)
        result_window.title("Weekly Scores")
        for name, score in results:
            ttk.Label(result_window, text=f"{name}: {score}").pack()

        
    def export_to_excel(self):
        conn = sqlite3.connect('tasks.db')
        
        today = datetime.now().date()
        days_until_thursday = (3 - today.weekday()) % 7
        thursday = today + timedelta(days=days_until_thursday)
        
        query = """SELECT date, task, name, A分數, B分數, C分數, YYYYMM, YYYY FROM WEEKTASK_REPORT WHERE date = ?"""
        df = pd.read_sql_query(query, conn, params=(thursday.strftime('%Y-%m-%d'),))
        conn.close()

        excel_path = "tasks_export.xlsx"
        sheet_name = "X"

        try:
            workbook = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
        else:
            sheet = workbook.create_sheet(sheet_name)
            last_row = 0

        # Write header if sheet is empty
        if last_row == 0:
            for col, header in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col, value=header)
            last_row = 1

        # Append new data
        for row, data in enumerate(df.values, start=last_row + 1):
            for col, value in enumerate(data, start=1):
                sheet.cell(row=row, column=col, value=value)

        workbook.save(excel_path)
        tk.messagebox.showinfo("Export Successful", f"Data has been appended to {sheet_name} sheet in tasks_export.xlsx")

if __name__ == "__main__":
    root = tk.Tk()
    app = TaskToolGUI(root)
    root.mainloop()
